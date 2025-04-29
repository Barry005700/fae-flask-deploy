from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from openpyxl import Workbook, load_workbook
import os
from datetime import datetime

app = Flask(__name__)
CORS(app)

def get_excel_filename():
    today = datetime.today()
    return f'fae_tasks_{today.year}-{today.month:02d}.xlsx'

@app.route('/')
def serve_index():
    return send_file('index.html')

@app.route('/submit-task', methods=['POST'])
def submit_task():
    data = request.json
    if not data:
        return jsonify({'error': 'No data provided'}), 400

    excel_file = get_excel_filename()

    if os.path.exists(excel_file):
        wb = load_workbook(excel_file)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(['日期', '部門', 'FAE', '產品線', '內容', '工時'])

    ws.append([
        data['date'],
        data['department'],
        data['fae'],
        data['product_line'],
        data['task_description'],
        data['hours']
    ])
    wb.save(excel_file)
    return jsonify({'message': '任務已儲存'}), 200

@app.route('/tasks/<fae>', methods=['GET'])
def get_tasks(fae):
    excel_file = get_excel_filename()
    if not os.path.exists(excel_file):
        return jsonify([])

    wb = load_workbook(excel_file)
    ws = wb.active
    tasks = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[2] == fae:
            tasks.append({
                'date': row[0],
                'department': row[1],
                'fae': row[2],
                'product_line': row[3],
                'task_description': row[4],
                'hours': row[5]
            })
    return jsonify(tasks)

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000)